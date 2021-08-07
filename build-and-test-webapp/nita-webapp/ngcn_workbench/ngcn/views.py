""" ********************************************************

Project: nita-webapp

Copyright (c) Juniper Networks, Inc., 2021. All rights reserved.

Notice and Disclaimer: This code is licensed to you under the Apache 2.0 License (the "License"). You may not use this code except in compliance with the License. This code is not an official Juniper product. You can obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0.html

SPDX-License-Identifier: Apache-2.0

Third-Party Code: This code may depend on other components under separate copyright notice and license terms. Your use of the source code for those components is subject to the terms and conditions of the respective license as noted in the Third-Party source code file.

******************************************************** """
import os
import json
import jenkins
import yaml
from yaml.constructor import Constructor
import collections
from collections import OrderedDict
from io import StringIO
import traceback
import logging
import configparser
import re

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.core import serializers
from django.utils import timezone
from django.utils.encoding import smart_str
from django.utils.translation import gettext as _
from django.core.files.storage import default_storage
from django.views.generic import View
from django.contrib.auth.decorators import login_required
from django.conf import settings

from openpyxl import load_workbook
from openpyxl import Workbook as open_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment
from xml.dom.minidom import parseString
from django_tables2   import RequestConfig
import django_tables2 as tables

from ngcn.models import CampusType,CampusNetwork,ActionHistory,ActionCategory,Workbook,Worksheets,Action,Resource,Role
from ngcn.tables import CampusTypeTable,CampusNetworkTable,ActionHistoryTable,ActionListTable, CampusNetworkActionListTable, RolesTable, ResourcesTable
from ngcn.forms import CampusNetworkForm, UploadFileForm, CampusTypeForm, EditCampusNetworkForm
from ngcn.networktypeparser import NetworkTypeParser
from ngcn.utils import ServerProperties
from ngcn.utils import wait_and_get_build_status
from docutils.parsers.rst import roles

TEST_ACTION_ID = 3

config = configparser.ConfigParser()
config_location = settings.BASE_DIR+"/../"
config.read_file(open(config_location + 'server_details.ini'))
jenkins_host_name = config['jenkins.server.details']['hostname']
jenkins_port =config['jenkins.server.details']['port']
JENKINS_SERVER_URL = 'http://' + jenkins_host_name + ':' + str(jenkins_port)
JENKINS_SERVER_USER=os.getenv('JENKINS_USER', "admin")
JENKINS_SERVER_PASS=os.getenv('JENKINS_PASS', "admin")
server = jenkins.Jenkins(JENKINS_SERVER_URL, username=JENKINS_SERVER_USER, password=JENKINS_SERVER_PASS)

logger=logging.getLogger(__name__)

@login_required(login_url='/admin/login/')
def treeView(request):

    campusTypeJson = {}
    campusTypeJson['name'] = _('network_type_heading') + 's'
    campusTypeJson['id'] = 'campus_type'
    campusTypeJson['children'] = []

    campusTypeList = CampusType.objects.all()
    for campusType in campusTypeList:
        lable = campusType.name
        type_id = 'campustype_'+ str(campusType.id)
        record = {'name':str(lable), 'id':str(type_id)}
        campusTypeJson['children'].append(record)

    campusNetJson = {}
    campusNetJson['name'] = _('network_heading') + 's'
    campusNetJson['id'] = 'campus_network'
    campusNetJson['children'] = []

    campusNetList = CampusNetwork.objects.all()
    for campusNet in campusNetList:
        lable = campusNet.name
        net_id = 'campusnetwork_'+ str(campusNet.id)
        record = {'name':str(lable), 'id':str(net_id)}
        campusNetJson['children'].append(record)

    treeJson = []
    treeJson.append(campusTypeJson)
    treeJson.append(campusNetJson)

    logger.info(treeJson)

    return render(request, 'ngcn/tree_view.html', {'data':treeJson})

def getTreeData(request):

    campusTypeJson = {}
    campusTypeJson['name'] = _('network_type_heading') + 's'
    campusTypeJson['id'] = 'campus_type'
    campusTypeJson['children'] = []

    campusTypeList = CampusType.objects.all()
    for campusType in campusTypeList:
        lable = campusType.name
        type_id = 'campustype_'+ str(campusType.id)
        record = {'name':str(lable), 'id':str(type_id)}
        campusTypeJson['children'].append(record)

    campusNetJson = {}
    campusNetJson['name'] = _('network_heading') + 's'
    campusNetJson['id'] = 'campus_network'
    campusNetJson['children'] = []

    campusNetList = CampusNetwork.objects.all()
    for campusNet in campusNetList:
        lable = campusNet.name
        net_id = 'campusnetwork_'+ str(campusNet.id)
        record = {'name':str(lable), 'id':str(net_id)}
        campusNetJson['children'].append(record)

    treeJson = []
    treeJson.append(campusTypeJson)
    treeJson.append(campusNetJson)

    logger.info(treeJson)

    return JsonResponse(treeJson, safe=False)

@login_required(login_url='/admin/login/')
def indexView(request):
    return render(request, 'ngcn/index.html')

def campusTypeMgmtView(request):
    logger.debug("campusTypeMgmtView: entered")
    queryset = CampusType.objects.all()
    table = CampusTypeTable(queryset)
    RequestConfig(request,paginate=False).configure(table)
    return render(request, 'ngcn/campus_type_mgmt.html',{'table':table})

def campusTypeView(request, campus_type_id):
    logger.debug("campusTypeView: entered")
    r = ""
    try:
        queryset = CampusType.objects.filter(id = campus_type_id)
        json_data = json.loads(serializers.serialize('json', queryset, fields=('name','app_zip_name','description')))
        form = json_data[0]["fields"]
        campus_networks = CampusNetwork.objects.filter(campus_type=campus_type_id)
        table = CampusNetworkTable(campus_networks)
        table.exclude = ('campus_type',)
        action_list_queryset = Action.objects.filter(campus_type_id = campus_type_id)
        action_list_table = ActionListTable(action_list_queryset)

        #roles_queryset = Role.objects.all().filter(campustype=CampusType.objects.get(id=campus_type_id))
        #roles_table = RolesTable(roles_queryset)

        #resources_queryset = Resource.objects.all().filter(campustype=CampusType.objects.get(id=campus_type_id))
        #resources_table = ResourcesTable(resources_queryset)
        RequestConfig(request,paginate=False).configure(table)

        logger.debug("campusTypeView: render the campus_type.html template")
        r = render(request, 'ngcn/campus_type.html',
                {
                    'form' :form,
                    'campus_network_table' :table,
                    'action_list_table' :action_list_table,
                    #'roles_table' :roles_table,
                    #'resources_table' :resources_table,
                    'campus_type_id' :campus_type_id})

        logger.debug("campusTypeView: " + r)
    except BaseException as e:
        logger.error("campusTypeView: render error {}".format(e))
    return r

def campusNetworkMgmtView(request):
    queryset = CampusNetwork.objects.all()
    table = CampusNetworkTable(queryset)
    RequestConfig(request,paginate=False).configure(table)
    logger.debug("campusNetworkMgmtView: exit")
    return render(request, 'ngcn/campus_network_mgmt.html',{'table':table})

def campusNetworkView(request, campus_network_id):
    return render(request, 'ngcn/campus_network.html',{'campus_network_id':campus_network_id})

class ConfigurationView(View):
    def get(self,request):
        return render(request,'ngcn/campus_network/upload_file.html',{})

def uploadFileView(request,campus_network_id):
    if request.method == 'POST':
        logger.debug("uploadFileView IF")
        try:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                result=parse_workbook(request.FILES['up_file'], campus_network_id)
                if result != "invalid_host" :
                    data_manager = GridDataManager()
                    sheets = data_manager.get_sheets_by_campus_network(campus_network_id)
                    table_json = json.dumps(sheets)
                    logger.debug(table_json)
                    return JsonResponse({'workbook':table_json,'status':'success'})
                else:
                    logger.error("invalid host in excel file")
                    return JsonResponse({'workbook':"",'status':'failure',"message":"Invalid data in host column in Excel file"})
            else:
                logger.info("invalid post request")
        except Exception as e:
            logger.error("Error while loading the Excel file.")
            logger.error(e)
            logger.error(traceback.print_exc())
            logger.error(traceback.format_exc())
            return JsonResponse({'workbook':"",'status':'failure',"message":"Error while loading the Excel file"})
    else:
        logger.debug("uploadFileView ELSE")
        try:
            data_manager = GridDataManager()
            sheets = data_manager.get_sheets_by_campus_network(campus_network_id)
            table_json = json.dumps(sheets)
            return JsonResponse({'workbook':table_json,'status':'success'})
        except Exception as e:
            logger.error("Error while parsing loading the data.")
            logger.error(e)
            logger.error(traceback.print_exc())
            logger.error(traceback.format_exc())
            return JsonResponse({'workbook':"",'status':'failure',"message":"Error while parsing loading the data"})

def configurationView(request,campus_network_id):
    return render(request, 'ngcn/campus_network/upload_file.html',{'campus_network_id':campus_network_id})

@login_required(login_url='/admin/login/')
def campusNetworkSummaryView(request,campus_network_id):
    queryset = CampusNetwork.objects.filter(id = campus_network_id)
    json_data = json.loads(serializers.serialize('json', queryset, fields=('name','status','description')))
    campus_network = CampusNetwork.objects.get(id = campus_network_id)
    campusTypeName = CampusType.objects.get(pk=campus_network.campus_type_id).name
    form = json_data[0]["fields"]
    form['ct_name']=campusTypeName;
#     print form
    action_list_queryset = Action.objects.filter(campus_type_id = campus_network.campus_type.id)
    action_list_table = CampusNetworkActionListTable(action_list_queryset, campus_network.name)
    return render(request, 'ngcn/campus_network/summary.html',{'form':form,'action_list_table' :action_list_table,'campus_network_id':campus_network_id})

@login_required(login_url='/admin/login/')
def actionsView(request,campus_network_id,action_category_id):
    category = ActionCategory.objects.get(pk=action_category_id)
    queryset = ActionHistory.objects.filter(campus_network_id = campus_network_id,category_id = category.id)
    campus_networks = CampusNetwork.objects.filter(id=campus_network_id)
    actions = None
    if campus_networks:
        campus_type_obj_id = campus_networks[0].campus_type.id
        actions = Action.objects.filter(action_category=category,campus_type_id=campus_type_obj_id)
    actionsJson = []
    for action_obj in actions:
        actionJson = {'name':action_obj.action_name,'url':action_obj.jenkins_url, 'id':action_obj.id}
        actionsJson.append(actionJson)
#     print actionsJson
    table = ActionHistoryTable(queryset)
    RequestConfig(request,paginate=False).configure(table)
    return render(request, 'ngcn/campus_network/noob.html',{'table':table, 'actions':actionsJson})

@login_required(login_url='/admin/login/')
def actionHistoryView(request,campus_network_id):
    queryset = ActionHistory.objects.filter(campus_network_id = campus_network_id).order_by('-timestamp')
    host = request.get_host()
    host_name = request.get_host().rsplit(':', 1)[0]
    jenkins_port = ServerProperties.getServerPort()
    host = host_name+":"+jenkins_port
    table = createActionHistory(queryset,host)
#    table = AllActionHistoryTable(queryset)
    RequestConfig(request,paginate=False).configure(table)
    return render(request, 'ngcn/campus_network/action_history.html',{'table':table})

@login_required(login_url='/admin/login/')
def actionHistoryDetailView(request,campus_network_id, action_category_id):
    queryset = ActionHistory.objects.filter(campus_network_id = campus_network_id, category_id=ActionCategory.objects.get(pk=action_category_id)).order_by('-timestamp')
    table = ActionHistoryTable(queryset)
    RequestConfig(request,paginate=False).configure(table)
    return render(request, 'ngcn/campus_network/campus_action_history.html',{'table':table})

@login_required(login_url='/admin/login/')
def saveGridDataView(request,campus_network_id):
    # grid data to the ordered dict
    logger.debug("grid data to the ordered dict")
    try:
        grid_list = json.JSONDecoder(object_pairs_hook=collections.OrderedDict).decode(request.body.decode("utf-8"))
        data_manager = GridDataManager()
        sheets = data_manager.get_sheets_by_campus_network(campus_network_id)
        for grid in grid_list['data'] :
            for sheet in sheets:
                if sheet["name"] == grid["name"]:
                    sheet[grid["name"]] = grid[grid["name"]]
        data_manager.create_or_update_db(campus_network_id,sheets)
        return HttpResponse("Success")
    except Exception as e:
        logger.error("Error while saving config talbe.")
        logger.error(e)
        logger.error(traceback.print_exc())
        logger.error(traceback.format_exc())
        return HttpResponse("Failure")

@login_required(login_url='/admin/login/')
def createExcelFileView(request,campus_network_id):
    try:
        create_workbook(campus_network_id)
        return HttpResponse("Success")
    except Exception as e:
        logger.error("Error while creating the workbook")
        logger.error(e)
        logger.error(traceback.print_exc())
        logger.error(traceback.format_exc())
        return HttpResponse("Failure")

@login_required(login_url='/admin/login/')
def downloadConfigDataView(request,campus_network_id):
    data_manager = GridDataManager()
    workbook_name = data_manager.get_workbook_name(campus_network_id)
    excel = open("/tmp/nita-webapp/export/"+workbook_name,"rb")
    output = excel.read()
    excel.close()
    response = HttpResponse(output,content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(workbook_name)
    logger.debug("downloadConfigDataView: exit")
    return response

@login_required(login_url='/admin/login/')
def clearGridDataView(request,campus_network_id):
    data_manager = GridDataManager()
    data_manager.delete_workbook(campus_network_id)
    return HttpResponse("Success")

@login_required(login_url='/admin/login/')
def jenkinsConsoleView(request,action_history_id):
    action_history = ActionHistory.objects.get(id=action_history_id)
    job_url = action_history.action_id.jenkins_url
    job_id = action_history.jenkins_job_build_no
    jenkins_port = ServerProperties.getServerPort()
    host = request.get_host()
    host_name = request.get_host().rsplit(':', 1)[0]
    console_url = "http://" + host_name + ":" + jenkins_port + "/job/" + job_url + "/" + str(job_id) + "/consoleText"
    logger.debug(console_url)
    return HttpResponse(console_url)

def escape_ansi(line):
    ansi_escape = re.compile(r'(?:\x1B[@-_]|[\x80-\x9F])[0-?]*[ -/]*[@-~]')
    return ansi_escape.sub('', line)

@login_required(login_url='/admin/login/')
def jenkinsConsoleLogView(request,action_history_id):
    action_history = ActionHistory.objects.get(id=action_history_id)
    logger.debug(action_history_id)
    logger.debug(action_history)
    campus_network = action_history.campus_network_id
    job_url = action_history.action_id.jenkins_url + "-" + campus_network.name
    logger.debug(job_url)
    job_id = action_history.jenkins_job_build_no
    logger.debug(job_id)
    try:
        response = server.get_build_console_output(job_url, job_id)
        #logger.debug(response)
    except:
        response = "The current build is queued in the Jenkins Server. Please wait for some time to view console log.. "
    return HttpResponse(escape_ansi(response))

''' forms '''

# def addCampusTypeView(request):
#     formset = ActionFormSet()
#     if request.method == 'POST':
#             form = CampusTypeForm(request.POST,request.FILES)
#             if form.is_valid():
#                 campusTypeForm  = form.save(commit=False)
#                 formset = ActionFormSet(request.POST,instance=campusTypeForm)
#                 if formset.is_valid():
#                     campusTypeForm.name=campusTypeForm.name.strip()
#                     campusTypeForm.description=campusTypeForm.description.strip()
#                     try:
#                         campusTypeForm.app_zip_name = request.FILES['app_zip_file'].name
#                     except:
#                         return JsonResponse({'result':'invalid_zip', 'name':campusTypeForm.name})
#                     result=validateZip(request.FILES['app_zip_file'],formset)
#                     if result['status'] == 'failed':
#                         return JsonResponse({'result':'invalid_zip', 'name':campusTypeForm.name})
#                     action_url = 'network_type_validator';
#                     #initiating the copy_template_job
#                     tree = parseString(server.get_job_config(action_url))
#                     tree.getElementsByTagName("customWorkspace")[0].firstChild.replaceWholeText(getWorkspaceLocation())
#                     server.reconfig_job(action_url, tree.toxml())
#                     current_build_number = server.get_job_info(action_url)['nextBuildNumber']
#                     app_zip = request.FILES['app_zip_file']
#                     with open(app_zip.temporary_file_path(),"r") as app_zip_file:
#                         Jenkins(JENKINS_SERVER_URL).get_job(action_url).invoke(
#                         build_params={
#                             'file_name': request.FILES['app_zip_file'].name,
#                             'operation':'add',
#                             'project_workspace_path': getWorkspaceLocation()
#                         },
#                         files={ 'app.zip':app_zip_file}
#                     )
#                     if wait_and_get_build_status(action_url,current_build_number) :
#                         campusTypeForm.save()
#                         formset.save()
#                         return JsonResponse({'result':'success', 'name':campusTypeForm.name})
#                     else:
#                         return JsonResponse({'result':'job_failed', 'name':campusTypeForm.name})
#                 else:
#                     return render(request, 'ngcn/campus_type/add.html', {'form': form,'formset':formset})
#             else:
#                 print "------- invalid request ------"
#                 return render(request, 'ngcn/campus_type/add.html', {'form': form,'formset':formset})
#
#     form = CampusTypeForm()
#     return render(request, 'ngcn/campus_type/add.html', {
#         'form': form,
#         'formset':formset
#     })

@login_required(login_url='/admin/login/')
def addCampusTypeView(request):
    if request.method == 'POST':
        form = CampusTypeForm(request.POST,request.FILES)
        if form.is_valid():
            zip_file = request.FILES['app_zip_file']
            networkTypeParser = NetworkTypeParser()
            if default_storage.exists(zip_file.name) :
                default_storage.delete(zip_file.name)
            file_name = default_storage.save(zip_file.name,zip_file)
            zip_file.close()
            data = networkTypeParser.parseProjectFile(file_name)
            default_storage.delete(file_name)
            logger.info(data)
            logger.debug(data)
            return data
        else:
            logger.info("Invalid form request")
            return render(request, 'ngcn/campus_type/add.html', {'form': form})

    form = CampusTypeForm()
    return render(request, 'ngcn/campus_type/add.html', {
        'form': form,
    })


'''
def editCampusTypeView(request, campus_type_id):
    formset = ActionFormSet()
    if request.method == 'POST':
            POST = request.POST.copy()
            campusType = CampusType.objects.get(pk=campus_type_id)
            POST['name']=campusType.name
            form = EditCampusTypeForm(POST,request.FILES,instance=campusType)
            if form.is_valid():
                campusTypeForm  = form.save(commit=False)
                formset = ActionFormSet(POST,request.FILES,instance=campusType)
                if formset.is_valid():
                    campusTypeForm.description=campusTypeForm.description.strip()
                    action_url = 'network_type_validator'
                    current_build_number = server.get_job_info(action_url)['nextBuildNumber']
                    try:
                        old_file_name = campusTypeForm.app_zip_name
                        app_zip = request.FILES['app_zip_file']
                        campusTypeForm.app_zip_name = app_zip.name
                        result=validateZip(request.FILES['app_zip_file'],formset)
                        if result['status'] == 'failed':
                            return JsonResponse({'result':'invalid_zip', 'name':campusTypeForm.name})
                        with open(app_zip.temporary_file_path(),"r") as app_zip_file:
                            Jenkins(JENKINS_SERVER_URL).get_job(action_url).invoke(
                                build_params={
                                    'old_file_name' : old_file_name,
                                    'file_name': app_zip.name,
                                    'operation':'edit'
                                },
                                files={ 'app.zip':app_zip_file}
                            )
                    except Exception, err :
                        print err
                        return JsonResponse({'result':'invalid_zip', 'name':campusTypeForm.name})
                        pass

                    if wait_and_get_build_status(action_url,current_build_number) :
                        campusTypeForm.save()
                        formset.save()
                        return JsonResponse({'result':'success', 'name':campusTypeForm.name})
                    else:
                        return JsonResponse({'result':'job_failed', 'name':campusTypeForm.name})
#                     campusTypeForm.save()
#                     formset.save()
#                     return JsonResponse({'result':'success', 'name':campusTypeForm.name})
                else:
                    return render(request, 'ngcn/campus_type/add.html', {'form': form,'formset':formset})
            else:
                print "------- invalid request ------"
                return render(request, 'ngcn/campus_type/add.html', {'form': form,})
    campusType = CampusType.objects.get(pk=campus_type_id)
    formset =  ActionFormSet(instance=campusType)
    form = EditCampusTypeForm(instance=campusType)
    return render(request, 'ngcn/campus_type/add.html', {
        'form': form,
        'formset':formset
    })
'''

@login_required(login_url='/admin/login/')
def deleteCampusTypeView(request):

    if request.method == 'POST':
        campus_type_id=request.POST.get('campus_type_ids')
#         print campus_type_ids
        campusType = CampusType.objects.get(pk=campus_type_id)
        name=campusType.name
        campusNetwork = CampusNetwork.objects.filter(campus_type=campusType)
        if campusNetwork:
            return JsonResponse({'status':'failed',
                                 'message':'The selected ' + _('network_type_heading') + ' has active ' + _('network_heading') + '(s). Please delete the active ' + _('network_heading') + 's before deleting the ' + _('network_type_heading'), 'name':name})

        action_url = 'network_type_validator';
        current_build_number = server.get_job_info(action_url)['nextBuildNumber']
        Jenkins(JENKINS_SERVER_URL, username=JENKINS_SERVER_USER, password=JENKINS_SERVER_PASS).get_job(action_url).invoke(build_params={'file_name': campusType.app_zip_name,
                                                        'operation':'delete',
                                                        'project_workspace_path': ServerProperties.getWorkspaceLocation()},
                                         )
        if wait_and_get_build_status(action_url,current_build_number) :
            CampusType.objects.get(pk=campusType.id).delete()
        return JsonResponse({'status':'success', 'name':name})

''' Campus Network forms '''

@login_required(login_url='/admin/login/')
def addCampusNetworkView(request):

    if request.method == 'POST':
        form = CampusNetworkForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            campusNetwork = CampusNetwork(name=data['name'].strip())
            campusNetwork.description=data['description'].strip();
            campusNetwork.dynamic_ansible_workspace=data['dynamic_ansible_workspace']
            #campusNetwork.status=data['status'].strip();
            campusNetwork.status = "Initialized";
            campusNetwork.host_file = data['host_file'].read().decode();
            campusNetwork.campus_type = data['campus_type'];
            src = ServerProperties.getWorkspaceLocation()+"/"+campusNetwork.campus_type.app_zip_name
            network_name = campusNetwork.name
            network_desc = campusNetwork.description
            action_url = 'network_template_mgr';
            #initiating the copy_template_job
            current_build_number = server.get_job_info(action_url)['nextBuildNumber']

            #generating the project.yaml file content
            project_yaml={}
            project_yaml['name']=network_name
            project_yaml['description']=campusNetwork.description
            project_yaml['action']=[]
            project_yaml['roles_and_resources']={}
            project_yaml['roles_and_resources']['roles']=[]
            project_yaml['roles_and_resources']['resources']=[]

            action_suffix= "("+network_name+")"

            jobs_list=Action.objects.filter(campus_type_id=campusNetwork.campus_type)
            for job in jobs_list:
                action={}
                action['name']=job.action_name+action_suffix
                action['jenkins_url']=job.jenkins_url+"-"+network_name
                action['category']=job.action_category.category_name

                job_property=job.action_property
                configuration={}
                configuration['shell_command']=job_property.shell_command
                configuration['output_path']=job_property.output_path
                configuration['custom_workspace']=job_property.custom_workspace

                action['configuration']=configuration
                project_yaml['action'].append(action)

                logger.debug('action[name] ' + str(action['name']))
                logger.debug('action[jenkins_url] ' + str(action['jenkins_url']))
                logger.debug('action[category] ' + str(action['category']))
                logger.debug('configuration[shell_command] ' + str(configuration['shell_command']))
                logger.debug('configuration[output_path] ' + str(configuration['output_path']))
                logger.debug('configuration[custom_workspace] ' + str(configuration['custom_workspace']))

            roles_list=Role.objects.all().filter(campustype=campusNetwork.campus_type)
            if roles_list:
                for role in roles_list:
                    project_yaml['roles_and_resources']['roles'].append(role.name)
            else:
                logger.info("roles for Campus Type: " + campusNetwork.campus_type.name + "::::: is Empty ")

            resources_list=Resource.objects.all().filter(campustype=campusNetwork.campus_type)
            if resources_list:
                for resource in resources_list:
                    project_yaml['roles_and_resources']['resources'].append(resource.name)
            else:
                logger.info("resources for Campus Type: " + campusNetwork.campus_type.name + "::::: is Empty ")

            #server.build_job(action_url,{'operation':'create', 'src':src, 'network_name':network_name, 'hosts':campusNetwork.host_file, 'project_yaml':yaml.safe_dump(project_yaml, default_flow_style=False)})
            server.build_job(action_url,{'operation':'create', 'src':src, 'network_name':network_name, 'hosts':campusNetwork.host_file, 'network_desc':network_desc})

            if wait_and_get_build_status(action_url,current_build_number) :
                logger.info("Campus Network - " + campusNetwork.name + " creation network_template_manager job finished successfully")
                campusNetwork.save()
                return JsonResponse({'result':'success', 'name':network_name})
            else:
                logger.error("Campus Network - " + campusNetwork.name + " creation network_template_manager job failed.")
                logger.error("Deleting the partially created Campus Network")

                current_build_number = server.get_job_info(action_url)['nextBuildNumber']
                server.build_job(action_url,{'operation':'delete', 'src':src, 'network_name':network_name})
                wait_and_get_build_status(action_url,current_build_number)
                return JsonResponse({'result':'failure', 'name':network_name})
        else:
            logger.info("------- invalid request ------")
            return render(request, 'ngcn/campus_network/add_campus_network.html', {'form': form,})

    form = CampusNetworkForm()
    return render(request, 'ngcn/campus_network/add_campus_network.html', {
        'form': form,
    })

@login_required(login_url='/admin/login/')
def editCampusNetworkView(request, campus_network_id):

    if request.method == 'POST':
            campusNetwork = CampusNetwork.objects.get(pk=campus_network_id)
            POST = request.POST.copy()
            POST['campus_type']=campusNetwork.campus_type.id
            form = EditCampusNetworkForm(data=POST, instance=campusNetwork)
            if form.is_valid():
                campusNetwork = CampusNetwork.objects.get(pk=campus_network_id)
                data = form.cleaned_data
                campusNetwork.description=data['description'].strip();
                campusNetwork.dynamic_ansible_workspace=data['dynamic_ansible_workspace'];
#                 campusNetwork.status=data['status'].strip();
                #campusNetwork.host_file=data['host_file'];
                hostData=data['host_file'];   #data['host_file'].read()
                src=ServerProperties.getWorkspaceLocation()+"/"+campusNetwork.campus_type.app_zip_name
                network_name=data['name'].strip()
                action_url = 'network_template_mgr';
                current_build_number = server.get_job_info(action_url)['nextBuildNumber']
                server.build_job(action_url,{'src':src,'network_name':network_name,'hosts':hostData,'operation':'update'})

                if wait_and_get_build_status(action_url,current_build_number) :
                    campusNetwork.status = "Hosts file modified"
                    campusNetwork.host_file=hostData
                    #print campusNetwork.host_file
                    campusNetwork.campus_type=data['campus_type'];
                    campusNetwork.save()
                    return JsonResponse({'result':'success', 'name':network_name})
                else:
                    campusNetwork.status = "Hosts file modification failed"
                    campusNetwork.save()
                    return JsonResponse({'result':'failure', 'name':network_name})
            else:
                logger.info("------- invalid request ------")
                return render(request, 'ngcn/campus_network/add_campus_network.html', {'form': form,})

    campusNetwork = CampusNetwork.objects.get(pk=campus_network_id)
    form = EditCampusNetworkForm(instance=campusNetwork)
    return render(request, 'ngcn/campus_network/add_campus_network.html', {
        'form': form,
    })

@login_required(login_url='/admin/login/')
def deleteCampusNetworkView(request):
    if request.method == 'POST':
        campus_network_ids=request.POST.get('campus_network_ids')

        campusNetwork = CampusNetwork.objects.get(pk=campus_network_ids)
        network_name=campusNetwork.name
        src=ServerProperties.getWorkspaceLocation()+"/"+campusNetwork.campus_type.app_zip_name

        action_url = 'network_template_mgr';
        current_build_number = server.get_job_info(action_url)['nextBuildNumber']
        server.build_job(action_url,{'src':src,'network_name':network_name,'operation':'delete'})
        if wait_and_get_build_status(action_url,current_build_number) :
            campusNetwork.delete()
            return JsonResponse({'status':'success', 'name':network_name})

        return JsonResponse({'status':'failure', 'name':network_name})

def parse_workbook(conf_file,campus_network_id):
    wb = load_workbook(conf_file)
    column_header = {}
    tables = []
    for sheet in wb :
        for row in sheet.iter_rows(min_row=2):
            if row[0].value is not None:
                var_dir = row[0].value.rpartition('/')[0].rpartition('/')[-1]
                if var_dir == 'group_vars' or var_dir == 'host_vars':
                    continue
                else:
                    return "invalid_host"
                    break

    for sheet in wb :
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    column_header[cell.column] = str(cell.value).replace(" ","_")
            break
        conf_value = []
        for row in sheet.iter_rows(min_row=2):
            conf = collections.OrderedDict()
            if row is not None and row[0].value is not None:
                for cell in row[:len(column_header)]:
                    conf[column_header[cell.column]] = cell.value
                conf_value.append(conf)
            else:
                break
        table = collections.OrderedDict()
        table["columns"] = build_column_data(conf_value[0].keys())
        table["name"] = sheet.title
        table[sheet.title] =    conf_value            #create_table_view(model, request)
        tables.append(table)
    #campus_network_name = request.session['app_name']
    try :
        workbook_row = Workbook.objects.get(campus_network_id=campus_network_id)
        if workbook_row is not None :
            Workbook.objects.get(campus_network_id=campus_network_id).delete()
    except:
        pass
    data_manager = GridDataManager()
    data_manager.create_or_update_db(campus_network_id,tables,conf_file.name)
    logger.debug("parse_workbook: exit")
    return "success"

def build_column_data(fields):
    columns =[]
    for field_value in fields:
        column = {}
        column["field"] = field_value
        column["id"] = field_value
        column["name"] = field_value
        columns.append(column)
    return columns

def ordered_dump(data, stream=None, Dumper=yaml.Dumper, **kwds):
    class OrderedDumper(Dumper):
        pass
    def _dict_representer(dumper, data):
        return dumper.represent_mapping(
            yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
            data.items())
    OrderedDumper.add_representer(OrderedDict, _dict_representer)
    OrderedDumper.add_representer(
        type(None),
        lambda dumper, value: dumper.represent_scalar(u'tag:yaml.org,2002:null', '')
    )
    return yaml.dump(data, stream, OrderedDumper, **kwds)

def add_bool(self, node):
    return self.construct_scalar(node)

Constructor.add_constructor(u'tag:yaml.org,2002:bool', add_bool)

def ordered_load(stream, Loader=yaml.Loader, object_pairs_hook=OrderedDict):
    class OrderedLoader(Loader):
        pass
    def construct_mapping(loader, node):
        loader.flatten_mapping(node)
        return object_pairs_hook(loader.construct_pairs(node))
    OrderedLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        construct_mapping)
    return yaml.load(stream, OrderedLoader)

def create_new_inv(workbook_name):
    workbook = load_workbook(workbook_name)
    print (workbook.sheetnames)

    from yamltoexcel import xls2yaml
    xls2yaml_instance = xls2yaml.ExcelToYaml(workbook_name,'./')
    for sheet_name in workbook.sheetnames:
        xls2yaml_instance.process_by_sheet(workbook, sheet_name)

    group_and_host_vars = OrderedDict()
    for host_file in xls2yaml_instance.sheet_data:
        #print "\n\nhost_file >>> ", host_file
        #base_yaml_content = yaml.safe_dump(xls2yaml_instance.sheet_data[host_file], default_flow_style=False, explicit_start=True)
        base_yaml_content=ordered_dump(OrderedDict(xls2yaml_instance.sheet_data[host_file]), Dumper=yaml.SafeDumper, default_flow_style=False, explicit_start=True)
        #print json.dumps(yaml.load(base_yaml_content))
        if "group_vars/" in host_file:
            host_file = host_file[host_file.find('group_vars/'):]
        elif "host_vars/" in host_file:
            host_file = host_file[host_file.find('host_vars/'):]
        else:
            return "Invalid host. The host must contain either /host_vars or /groups_vars in the path"
        group_and_host_vars.update( { host_file : ordered_load(base_yaml_content) })
    return group_and_host_vars


def create_workbook_from_db(campus_network_id):
    data_manager = GridDataManager()
    sheets = data_manager.get_sheets_by_campus_network(campus_network_id)

    wb = open_workbook()
    if "Sheet" in wb.sheetnames:
    	del wb["Sheet"]
    # create the sheet in workbook for each sheets from db

    sheet_index = 0
    for sheet in sheets:
        wb.create_sheet(index=sheet_index, title=sheet["name"])
        sheet_index+=1
    # populate the cells with db data for each sheet in workbook
    for sheet in sheets:
        ws = wb[sheet["name"]]
        keys = sheet[sheet["name"]][0].keys()
        column_header = {}
        # create the cells in workbook
        for i in range(1,len(sheet[sheet["name"]])+1):
            for j in range(1,len(keys)+1):
                ws.cell(row=i,column=j)

        column_index = 1
        # insert the column column_header value
        for key,cell in zip(keys,ws.iter_cols(min_row=1, max_row=1)) :
            cell[0].value = key
            column_header[cell[0].column] = key
            column_index+=1
        # insert the column data values
        row_index=2
        for row_sheet in sheet[sheet["name"]]:
            for column_index in range(1,len(keys)+1):
                ws.cell(row=row_index,column=column_index).value = row_sheet[column_header[ws.cell(row=row_index,column=column_index).column]]
            row_index+=1
    directory = '/tmp/nita-webapp/'
    if not os.path.exists(directory):
        os.makedirs(directory)
    workbook_name = '/tmp/nita-webapp/temp.xlsx'
    wb.save(workbook_name)
    return workbook_name


def create_workbook(campus_network_id):
    data_manager = GridDataManager()
    workbook_name = data_manager.get_workbook_name(campus_network_id)
    create_workbook_from_db(campus_network_id)
    workbook = load_workbook('/tmp/nita-webapp/temp.xlsx')
    print (workbook.sheetnames)

    from yamltoexcel import xls2yaml,yaml2xls
    xls2yaml_instance = xls2yaml.ExcelToYaml(workbook_name,'./')
    yaml2xls_instance = yaml2xls.YamlToExcel("")

    for sheet_name in workbook.sheetnames:
        xls2yaml_instance.process_by_sheet(workbook, sheet_name)

    #group_and_host_vars = OrderedDict()

    #wb = load_workbook()
    wb = open_workbook()
    ws = wb.active
    ws.title = 'base'

    # Styling
    value_font = Font(name="Bitstream Charter", size=10)
    vthin = Side(border_style="thin", color="000000")
    vborder = Border(top=vthin, left=vthin, right=vthin, bottom=vthin)
    valignment = Alignment(wrap_text=True)

    value_style = NamedStyle(name = "value", font = value_font, border = vborder, alignment = valignment)
    wb.add_named_style(value_style)

    header_font = Font(name="Bitstream Charter", size=10, bold=True)
    hthin = Side(border_style="thin", color="000000")
    hfill = PatternFill(fill_type='solid', fgColor="33bbff")
    hborder = Border(top=hthin, left=hthin, right=hthin, bottom=hthin)

    header_style = NamedStyle(name = "header", font = header_font, fill = hfill, border = hborder)
    wb.add_named_style(header_style)

    ws.cell(row=1, column=1).value = "host"
    ws.cell(row=1, column=1).style = "header"
    ws.cell(row=1, column=2).value = "name"
    ws.cell(row=1, column=2).style = "header"
    ws.cell(row=1, column=3).value = "value"
    ws.cell(row=1, column=3).style = "header"

    sheet_last_row_index = {}

    sheet_last_row_index['base'] = 1

    yaml2xls_instance.put_border(wb)
    for host_file in xls2yaml_instance.sheet_data:
        base_yaml_content =  OrderedDict(xls2yaml_instance.sheet_data[host_file])
        #base_yaml_content=ordered_dump(OrderedDict(xls2yaml_instance.sheet_data[host_file]), Dumper=yaml.SafeDumper, default_flow_style=False, explicit_start=True)
        host_name = host_file
        yaml2xls_instance.parse_yaml_files(wb, ws, base_yaml_content, host_name, sheet_last_row_index)

    directory = '/tmp/nita-webapp/export'
    if not os.path.exists(directory):
        os.makedirs(directory)

    wb.save('/tmp/nita-webapp/export/'+workbook_name)


# trigger job code

from jenkinsapi.jenkins import Jenkins
#new_server = Jenkins('http://jenkins:8080')

@login_required(login_url='/admin/login/')
def triggerAction(request, action_id, campus_network_id):
    action_obj = Action.objects.get(pk=action_id)
    action_url = action_obj.jenkins_url
    try:
        if Workbook.objects.filter(campus_network_id=campus_network_id).count() == 0:
            return JsonResponse({'status':'failure', 'name':action_obj.action_name, 'reason':'No data configured'})

        campus_network=CampusNetwork.objects.get(pk=campus_network_id)
        action_url+="-"+campus_network.name
        updateCampusNetworkStatusOnDB(campus_network_id, "Triggered " + action_url)
        campus_type = CampusType.objects.get(pk=campus_network.campus_type.id)
        workbook_name = create_workbook_from_db(campus_network_id)
        configuration_data = create_new_inv(workbook_name)
        logger.debug('configuration_data: ' + str(configuration_data))

        if configuration_data != "invalid file name" :
            if campus_network.dynamic_ansible_workspace is True:
                build_dir = "/var/tmp/build/" + campus_type.name + "-" + campus_network.name
            else:
                build_dir = configuration_data['group_vars/all.yaml']['build_dir']

            current_build_number = server.get_job_info(action_url)['nextBuildNumber']
            Jenkins(JENKINS_SERVER_URL, username=JENKINS_SERVER_USER, password=JENKINS_SERVER_PASS).get_job(action_url).invoke(
                files={
                    'data.json':json.dumps(configuration_data)
            },
            build_params={
                        'build_dir':build_dir
                    }
            )
        history = ActionHistory(action_id=Action.objects.get(pk=action_id), timestamp=timezone.now(), status="Running", jenkins_job_build_no=current_build_number, category_id=ActionCategory.objects.get(pk=action_obj.action_category.id), campus_network_id=CampusNetwork.objects.get(pk=campus_network_id))
        history.save()
        logger.info(history.id)
    except Exception as e:
        logger.error("received error")
        traceback.print_exc()
        logger.error(e)
        return JsonResponse({'status':'failure', 'name':action_obj.action_name})
    return JsonResponse({'status':'success', 'name':action_obj.action_name})


def updateBuildStatusOnDB(action_history_id, buildStatus):
    historyObj = ActionHistory.objects.get(pk=action_history_id)
    historyObj.status=buildStatus.title()
    historyObj.save()

class GridDataManager:
    # create or update woorkbook data
    def create_or_update_db(self,campus_network_id,sheets,filename=None):
        print("")
        campus_network = CampusNetwork.objects.get(id=campus_network_id)
        try:
            if filename is not None:
                workbook_row = Workbook.objects.get(name=filename,campus_network_id=campus_network_id)
            else:
                workbook_row = Workbook.objects.get(campus_network_id=campus_network_id)
        except:
            workbook_row = Workbook(name=filename,campus_network_id=campus_network)
            workbook_row.save()
        for sheet in sheets:
            try:
                worksheet_row = Worksheets.objects.get(name = sheet["name"],workbook_id = workbook_row)
                worksheet_row.data= sheet
                worksheet_row.save()
            except:
                worksheet_row = Worksheets(name = sheet["name"],data=sheet,workbook_id = workbook_row)
                worksheet_row.save()

    # get the sheets by campus network
    def get_sheets_by_campus_network(self,campus_network_id):
        workbook_row = Workbook.objects.filter(campus_network_id_id=campus_network_id)
        sheets = Worksheets.objects.filter(workbook_id__in=workbook_row).values('data').order_by('id')
        sheet_data = []
        for sheet in sheets :
            sheet_data.append(eval(sheet['data']))
        return sheet_data

    def delete_workbook(self,campus_network_id):
        workbook = Workbook.objects.get(campus_network_id=campus_network_id)
        status = workbook.delete()

    """ workbook name defined by user while uploading """
    def get_user_defined_workbook_name(self,campus_network_id):
        workbook = Workbook.objects.get(campus_network_id=campus_network_id)
        return workbook.name

    """ workbook name comprised of campus_type name and network name"""
    def get_workbook_name(self,campus_network_id):
        campus_network = CampusNetwork.objects.get(id=campus_network_id)
        workbook_name =  campus_network.campus_type.name+"_"+campus_network.name +".xlsx"
        return workbook_name

def createActionHistory(queryset,host):
    class AllActionHistoryTable(tables.Table):
        action_id = tables.Column(verbose_name="Name",accessor='action_id.action_name')
        category_name = tables.Column(accessor='category_id.category_name')
        #description  = tables.Column(verbose_name="Description")
        timestamp = tables.Column(verbose_name="Timestamp")
        status = tables.Column(verbose_name="Status")
        jenkins_job_build_no  = tables.Column(verbose_name="Jenkins Job Id") #tables.TemplateColumn('<a href="http://'+host+'/job/{{record.action_id.jenkins_url}}/{{record.jenkins_job_build_no}}/console" target="_blank" ">{{record.jenkins_job_build_no}}</a>')
        console_log_link  = tables.TemplateColumn('<a href="https://' + os.getenv('JENKINS_URL', "localhost") + ':' + os.getenv('JENKINS_PORT', "8443") + '/job/{{record.action_id.jenkins_url}}-{{record.campus_network_id.name}}/{{record.jenkins_job_build_no}}/console" target="_blank" ">View</a>')
        class Meta:
            model = ActionHistory
            attrs = {"class": "table table-condensed table-bordered","id":"all-action-history-table",'width':'98%'}
            orderable = False
            exclude = ('id','category_id','campus_network_id')
            sequence = ('action_id','category_name','timestamp','status','jenkins_job_build_no')

    table = AllActionHistoryTable(queryset)
    return table

def updateCampusNetworkStatusOnDB(campus_network_id,status):
    campus_network = CampusNetwork.objects.get(id=campus_network_id)
    campus_network.status = status
    campus_network.save()
