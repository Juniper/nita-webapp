# Copyright (c) Hewlett Packard Enterprise, 2026. All rights reserved.
# SPDX-License-Identifier: Apache-2.0

"""Workbook / spreadsheet business logic shared by the API layer.

Contains the workbook parsing, generation and persistence helpers plus the
``GridDataManager`` that were previously defined in the legacy ``ngcn.views``
module. Extracted so ``ngcn.api`` no longer imports from the server-rendered UI
code.
"""

import collections
import logging
import os
import re
from collections import OrderedDict

import yaml
from openpyxl import Workbook as open_workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from yaml.constructor import Constructor

from ngcn.models import CampusNetwork, Workbook, Worksheets

logger = logging.getLogger(__name__)


def escape_ansi(line):
    ansi_escape = re.compile(r"(?:\x1B[@-_]|[\x80-\x9F])[0-?]*[ -/]*[@-~]")
    return ansi_escape.sub("", line)


def build_column_data(fields):
    columns = []
    for field_value in fields:
        column = {}
        column["field"] = field_value
        column["id"] = field_value
        column["name"] = field_value
        columns.append(column)
    return columns


def ordered_dump(data, stream=None, Dumper=yaml.Dumper, **kwds):
    class OrderedDumper(Dumper):
        pass

    def _dict_representer(dumper, data):
        return dumper.represent_mapping(
            yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG, data.items()
        )

    OrderedDumper.add_representer(OrderedDict, _dict_representer)
    OrderedDumper.add_representer(
        type(None),
        lambda dumper, value: dumper.represent_scalar("tag:yaml.org,2002:null", ""),
    )
    return yaml.dump(data, stream, OrderedDumper, **kwds)


def add_bool(self, node):
    return self.construct_scalar(node)


Constructor.add_constructor("tag:yaml.org,2002:bool", add_bool)


def ordered_load(stream, Loader=yaml.Loader, object_pairs_hook=OrderedDict):
    class OrderedLoader(Loader):
        pass

    def construct_mapping(loader, node):
        loader.flatten_mapping(node)
        return object_pairs_hook(loader.construct_pairs(node))

    OrderedLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG, construct_mapping
    )
    return yaml.load(stream, OrderedLoader)


def parse_workbook(conf_file, campus_network_id):
    wb = load_workbook(conf_file)
    column_header = {}
    tables = []
    for sheet in wb:
        for row in sheet.iter_rows(min_row=2):
            if row[0].value is not None:
                var_dir = row[0].value.rpartition("/")[0].rpartition("/")[-1]
                if var_dir == "group_vars" or var_dir == "host_vars":
                    continue
                else:
                    return "invalid_host"
                    break

    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    column_header[cell.column] = str(cell.value).replace(" ", "_")
            break
        conf_value = []
        for row in sheet.iter_rows(min_row=2):
            conf = collections.OrderedDict()
            if row is not None and row[0].value is not None:
                for cell in row[: len(column_header)]:
                    conf[column_header[cell.column]] = cell.value
                conf_value.append(conf)
            else:
                break
        table = collections.OrderedDict()
        table["columns"] = build_column_data(conf_value[0].keys())
        table["name"] = sheet.title
        table[sheet.title] = conf_value  # create_table_view(model, request)
        tables.append(table)
    # campus_network_name = request.session['app_name']
    Workbook.objects.filter(campus_network_id=campus_network_id).delete()
    data_manager = GridDataManager()
    data_manager.create_or_update_db(campus_network_id, tables, conf_file.name)
    logger.debug("parse_workbook: exit")
    return "success"


def create_new_inv(workbook_name):
    workbook = load_workbook(workbook_name)
    print(workbook.sheetnames)

    from yamltoexcel import xls2yaml

    xls2yaml_instance = xls2yaml.ExcelToYaml(workbook_name, "./")
    for sheet_name in workbook.sheetnames:
        xls2yaml_instance.process_by_sheet(workbook, sheet_name)

    group_and_host_vars = OrderedDict()
    for host_file in xls2yaml_instance.sheet_data:
        # print "\n\nhost_file >>> ", host_file
        # base_yaml_content = yaml.safe_dump(xls2yaml_instance.sheet_data[host_file], default_flow_style=False, explicit_start=True)
        base_yaml_content = ordered_dump(
            OrderedDict(xls2yaml_instance.sheet_data[host_file]),
            Dumper=yaml.SafeDumper,
            default_flow_style=False,
            explicit_start=True,
        )
        # print json.dumps(yaml.load(base_yaml_content))
        if "group_vars/" in host_file:
            host_file = host_file[host_file.find("group_vars/") :]
        elif "host_vars/" in host_file:
            host_file = host_file[host_file.find("host_vars/") :]
        else:
            return "Invalid host. The host must contain either /host_vars or /groups_vars in the path"
        group_and_host_vars.update({host_file: ordered_load(base_yaml_content)})
    return group_and_host_vars


def create_workbook_from_db(campus_network_id):
    data_manager = GridDataManager()
    sheets = data_manager.get_sheets_by_campus_network(campus_network_id)

    wb = open_workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    # create the sheet in workbook for each sheets from db

    sheet_index = 0
    for sheet in sheets:
        wb.create_sheet(index=sheet_index, title=sheet["name"])
        sheet_index += 1
    # populate the cells with db data for each sheet in workbook
    for sheet in sheets:
        ws = wb[sheet["name"]]
        keys = sheet[sheet["name"]][0].keys()
        column_header = {}
        # create the cells in workbook
        for i in range(1, len(sheet[sheet["name"]]) + 1):
            for j in range(1, len(keys) + 1):
                ws.cell(row=i, column=j)

        column_index = 1
        # insert the column column_header value
        for key, cell in zip(keys, ws.iter_cols(min_row=1, max_row=1), strict=False):
            cell[0].value = key
            column_header[cell[0].column] = key
            column_index += 1
        # insert the column data values
        row_index = 2
        for row_sheet in sheet[sheet["name"]]:
            for column_index in range(1, len(keys) + 1):
                ws.cell(row=row_index, column=column_index).value = row_sheet[
                    column_header[ws.cell(row=row_index, column=column_index).column]
                ]
            row_index += 1
    directory = "/tmp/nita-webapp/"
    if not os.path.exists(directory):
        os.makedirs(directory)
    workbook_name = "/tmp/nita-webapp/temp.xlsx"
    wb.save(workbook_name)
    return workbook_name


def create_workbook(campus_network_id):
    data_manager = GridDataManager()
    workbook_name = data_manager.get_workbook_name(campus_network_id)
    create_workbook_from_db(campus_network_id)
    workbook = load_workbook("/tmp/nita-webapp/temp.xlsx")
    print(workbook.sheetnames)

    from yamltoexcel import xls2yaml, yaml2xls

    xls2yaml_instance = xls2yaml.ExcelToYaml(workbook_name, "./")
    yaml2xls_instance = yaml2xls.YamlToExcel("")

    for sheet_name in workbook.sheetnames:
        xls2yaml_instance.process_by_sheet(workbook, sheet_name)

    # group_and_host_vars = OrderedDict()

    # wb = load_workbook()
    wb = open_workbook()
    ws = wb.active
    ws.title = "base"

    # Styling
    value_font = Font(name="Bitstream Charter", size=10)
    vthin = Side(border_style="thin", color="000000")
    vborder = Border(top=vthin, left=vthin, right=vthin, bottom=vthin)
    valignment = Alignment(wrap_text=True)

    value_style = NamedStyle(
        name="value", font=value_font, border=vborder, alignment=valignment
    )
    wb.add_named_style(value_style)

    header_font = Font(name="Bitstream Charter", size=10, bold=True)
    hthin = Side(border_style="thin", color="000000")
    hfill = PatternFill(fill_type="solid", fgColor="33bbff")
    hborder = Border(top=hthin, left=hthin, right=hthin, bottom=hthin)

    header_style = NamedStyle(
        name="header", font=header_font, fill=hfill, border=hborder
    )
    wb.add_named_style(header_style)

    ws.cell(row=1, column=1).value = "host"
    ws.cell(row=1, column=1).style = "header"
    ws.cell(row=1, column=2).value = "name"
    ws.cell(row=1, column=2).style = "header"
    ws.cell(row=1, column=3).value = "value"
    ws.cell(row=1, column=3).style = "header"

    sheet_last_row_index = {}

    sheet_last_row_index["base"] = 1

    yaml2xls_instance.put_border(wb)
    for host_file in xls2yaml_instance.sheet_data:
        base_yaml_content = OrderedDict(xls2yaml_instance.sheet_data[host_file])
        # base_yaml_content=ordered_dump(OrderedDict(xls2yaml_instance.sheet_data[host_file]), Dumper=yaml.SafeDumper, default_flow_style=False, explicit_start=True)
        host_name = host_file
        yaml2xls_instance.parse_yaml_files(
            wb, ws, base_yaml_content, host_name, sheet_last_row_index
        )

    directory = "/tmp/nita-webapp/export"
    if not os.path.exists(directory):
        os.makedirs(directory)

    wb.save("/tmp/nita-webapp/export/" + workbook_name)


class GridDataManager:
    # create or update woorkbook data
    def create_or_update_db(self, campus_network_id, sheets, filename=None):
        print("")
        campus_network = CampusNetwork.objects.get(id=campus_network_id)
        try:
            if filename is not None:
                workbook_row = Workbook.objects.get(
                    name=filename, campus_network_id=campus_network_id
                )
            else:
                workbook_row = Workbook.objects.get(campus_network_id=campus_network_id)
        except Exception:
            workbook_row = Workbook(name=filename, campus_network_id=campus_network)
            workbook_row.save()
        for sheet in sheets:
            try:
                worksheet_row = Worksheets.objects.get(
                    name=sheet["name"], workbook_id=workbook_row
                )
                worksheet_row.data = sheet
                worksheet_row.save()
            except Exception:
                worksheet_row = Worksheets(
                    name=sheet["name"], data=sheet, workbook_id=workbook_row
                )
                worksheet_row.save()

    # get the sheets by campus network
    def get_sheets_by_campus_network(self, campus_network_id):
        workbook_row = Workbook.objects.filter(campus_network_id_id=campus_network_id)
        sheets = (
            Worksheets.objects.filter(workbook_id__in=workbook_row)
            .values("data")
            .order_by("id")
        )
        sheet_data = []
        for sheet in sheets:
            sheet_data.append(eval(sheet["data"]))
        return sheet_data

    def delete_workbook(self, campus_network_id):
        workbook = Workbook.objects.get(campus_network_id=campus_network_id)
        workbook.delete()

    """ workbook name defined by user while uploading """

    def get_user_defined_workbook_name(self, campus_network_id):
        workbook = Workbook.objects.get(campus_network_id=campus_network_id)
        return workbook.name

    """ workbook name comprised of campus_type name and network name"""

    def get_workbook_name(self, campus_network_id):
        campus_network = CampusNetwork.objects.get(id=campus_network_id)
        workbook_name = (
            campus_network.campus_type.name + "_" + campus_network.name + ".xlsx"
        )
        return workbook_name


def updateCampusNetworkStatusOnDB(campus_network_id, status):
    campus_network = CampusNetwork.objects.get(id=campus_network_id)
    campus_network.status = status
    campus_network.save()
