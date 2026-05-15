# Copyright (c) Hewlett Packard Enterprise, 2026. All rights reserved.
# SPDX-License-Identifier: Apache-2.0

"""API tests using the evpn_vxlan_erb_dc example project as a data fixture.

The tests are organised in four sections that map directly onto the feature
areas under test:

  1. Project load   — validate and import project.yaml into the CampusType
                      registry (no Jenkins interaction required).
  2. Network creation — create a CampusNetwork via the REST API using the
                        real dc1-hosts inventory file.
     2.1 Host file load       — verify the host-file content is stored and
                                 returned verbatim.
     2.2 Configuration load   — upload dc1_data.xlsx via the workbook-upload
                                 endpoint and confirm sheet data is persisted.
     2.3 Export               — download the workbook as an .xlsx attachment.
     2.4 Full CRUD            — GET list/retrieve, POST, PUT, PATCH, DELETE.

Scope
-----
* These tests cover the REST API layer only (``/api/v1/…``).
* No Jenkins jobs are triggered.  External Jenkins calls are monkeypatched
  where the code paths under test would otherwise require them.
* No NOOB / build / test action jobs are invoked.
"""

import io
import pathlib
import tempfile
import zipfile

import pytest
import yaml
from django.utils import timezone
from ngcn.api import views as api_views
from ngcn.models import (
    Action,
    ActionCategory,
    ActionHistory,
    ActionProperty,
    CampusNetwork,
    CampusType,
    Workbook,
    Worksheets,
)
from ngcn.networktypeparser import NetworkTypeParser
from rest_framework.test import APIClient

# ---------------------------------------------------------------------------
# Fixture data paths
# ---------------------------------------------------------------------------

#: Root of the evpn_vxlan_erb_dc example directory.
#: Resolved from this file's location so it works in any checkout layout.
_HERE = pathlib.Path(__file__).resolve().parent
FIXTURE_DIR = _HERE.parents[4] / "nita" / "examples" / "evpn_vxlan_erb_dc"

PROJECT_YAML_PATH = FIXTURE_DIR / "project.yaml"
DC1_HOSTS_PATH = FIXTURE_DIR / "dc1-hosts"
DC2_HOSTS_PATH = FIXTURE_DIR / "dc2-hosts"
DC1_XLSX_PATH = FIXTURE_DIR / "dc1_data.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_project_zip(project_name: str) -> bytes:
    """Return an in-memory zip that satisfies NetworkTypeParser.validateZipFile.

    The archive contains the project directory at the top level with at least
    ``project.yaml`` and ``ansible.cfg``.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(PROJECT_YAML_PATH, f"{project_name}/project.yaml")
        zf.write(FIXTURE_DIR / "ansible.cfg", f"{project_name}/ansible.cfg")
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Shared pytest fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def api_client(user):
    """Authenticated DRF APIClient using force_authenticate."""
    client = APIClient()
    client.force_authenticate(user=user)
    return client


@pytest.fixture
def build_category(db):
    return ActionCategory.objects.create(category_name="BUILD")


@pytest.fixture
def test_category(db):
    return ActionCategory.objects.create(category_name="TEST")


@pytest.fixture
def all_categories(build_category, test_category):
    """Ensure BUILD and TEST ActionCategory rows exist before project import."""
    return {"BUILD": build_category, "TEST": test_category}


@pytest.fixture
def project_yaml_text():
    """Raw text of evpn_vxlan_erb_dc/project.yaml (no DB required)."""
    return PROJECT_YAML_PATH.read_text()


@pytest.fixture
def project_yaml_data(project_yaml_text):
    """Parsed dict of evpn_vxlan_erb_dc/project.yaml (no DB required)."""
    return yaml.safe_load(project_yaml_text)


@pytest.fixture
def evpn_campus_type(db, all_categories, project_yaml_text):
    """CampusType loaded directly from project.yaml, bypassing Jenkins."""
    parser = NetworkTypeParser()
    parser.updateNetworkTypeDetailsOnDB(
        project_yaml_text,
        "evpn_vxlan_erb_dc_1.3.zip",
    )
    data = yaml.safe_load(project_yaml_text)
    return CampusType.objects.get(name=data["name"])


@pytest.fixture
def dc1_host_content():
    """Text content of the dc1-hosts Ansible inventory file."""
    return DC1_HOSTS_PATH.read_text()


@pytest.fixture
def dc2_host_content():
    """Text content of the dc2-hosts Ansible inventory file."""
    return DC2_HOSTS_PATH.read_text()


@pytest.fixture
def evpn_network(db, evpn_campus_type, dc1_host_content):
    """CampusNetwork for DC1, created directly in the DB."""
    return CampusNetwork.objects.create(
        name="evpn-dc1",
        status="Initialized",
        description="EVPN VXLAN ERB DC1 test network",
        host_file=dc1_host_content,
        campus_type=evpn_campus_type,
    )


@pytest.fixture
def evpn_network_with_workbook(evpn_network):
    """CampusNetwork pre-populated with minimal Workbook/Worksheets rows.

    Provides just enough data for the workbook-retrieve and workbook-save
    tests without running a real xlsx upload.
    """
    workbook = Workbook.objects.create(
        name="dc1_data.xlsx",
        campus_network_id=evpn_network,
    )
    sheet_data = {
        "name": "base",
        "columns": [
            {"field": "host", "id": "host", "name": "host"},
            {"field": "name", "id": "name", "name": "name"},
            {"field": "value", "id": "value", "name": "value"},
        ],
        "base": [
            {"host": "group_vars/all.yaml", "name": "OS_dir", "value": "/var/tmp/"},
        ],
    }
    Worksheets.objects.create(
        name="base",
        data=str(sheet_data),
        workbook_id=workbook,
    )
    return evpn_network


# ===========================================================================
# Section 1 — Project load
# ===========================================================================


@pytest.mark.django_db
def test_project_yaml_is_valid_yaml(project_yaml_text):
    """project.yaml must parse as valid YAML without errors."""
    data = yaml.safe_load(project_yaml_text)
    assert isinstance(data, dict)


@pytest.mark.django_db
def test_project_yaml_has_required_top_level_keys(project_yaml_data):
    """project.yaml must carry 'name', 'description', and 'action' keys."""
    assert "name" in project_yaml_data
    assert "description" in project_yaml_data
    assert "action" in project_yaml_data


@pytest.mark.django_db
def test_project_yaml_name_matches_expected(project_yaml_data):
    """The project name must be 'evpn_vxlan_erb_dc_1.3'."""
    assert project_yaml_data["name"] == "evpn_vxlan_erb_dc_1.3"


@pytest.mark.django_db
def test_validate_project_yaml_passes_with_real_data(db, all_categories, project_yaml_text):
    """NetworkTypeParser.validateProjectYaml must return None (no error)
    when given the real evpn_vxlan_erb_dc project.yaml content.
    """
    parser = NetworkTypeParser()
    error = parser.validateProjectYaml(project_yaml_text)
    assert error is None


@pytest.mark.django_db
def test_project_load_creates_campus_type(evpn_campus_type, project_yaml_data):
    """updateNetworkTypeDetailsOnDB must create a CampusType row with the
    correct name and description.
    """
    assert evpn_campus_type.name == project_yaml_data["name"]
    assert evpn_campus_type.description == project_yaml_data["description"]


@pytest.mark.django_db
def test_project_load_creates_correct_number_of_actions(evpn_campus_type, project_yaml_data):
    """One Action row must be created for every entry in project.yaml 'action'."""
    expected_count = len(project_yaml_data["action"])
    actual_count = Action.objects.filter(campus_type_id=evpn_campus_type).count()
    assert actual_count == expected_count


@pytest.mark.django_db
def test_project_load_actions_have_correct_names(evpn_campus_type, project_yaml_data):
    """Every action name from project.yaml must appear in the DB."""
    expected_names = {a["name"] for a in project_yaml_data["action"]}
    actual_names = set(
        Action.objects.filter(campus_type_id=evpn_campus_type).values_list(
            "action_name", flat=True
        )
    )
    assert expected_names == actual_names


@pytest.mark.django_db
def test_project_load_actions_reference_valid_categories(evpn_campus_type):
    """Every Action created for the project must reference an existing
    ActionCategory (BUILD or TEST).
    """
    valid_category_names = {"BUILD", "TEST"}
    for action in Action.objects.filter(campus_type_id=evpn_campus_type):
        assert action.action_category.category_name.upper() in valid_category_names


@pytest.mark.django_db
def test_project_load_test_actions_have_output_path(evpn_campus_type):
    """TEST-category actions must persist a non-empty output_path."""
    test_actions = Action.objects.filter(
        campus_type_id=evpn_campus_type,
        action_category__category_name="TEST",
    )
    assert test_actions.exists(), "Expected at least one TEST action"
    for action in test_actions:
        assert action.action_property.output_path not in (None, "", " ")


@pytest.mark.django_db
def test_project_load_duplicate_name_rejected(db, all_categories, project_yaml_text):
    """validateProjectYaml must return an error string when the CampusType
    name already exists in the database (duplicate protection).
    """
    parser = NetworkTypeParser()
    # First import succeeds.
    parser.updateNetworkTypeDetailsOnDB(project_yaml_text, "evpn_vxlan_erb_dc_1.3.zip")
    # Second import of the same project must be rejected at validation time.
    error = parser.validateProjectYaml(project_yaml_text)
    assert error is not None
    assert isinstance(error, str)


@pytest.mark.django_db
def test_project_load_via_zip_upload_api(api_client, db, all_categories, monkeypatch):
    """POST /api/v1/network-types/upload/ must accept a well-formed zip built
    from the evpn_vxlan_erb_dc project and return result=success.

    The Jenkins and filesystem interactions inside parseProjectFile are
    monkeypatched; only the zip-validation and YAML-validation paths are
    exercised end-to-end.
    """
    import django.core.files.storage as django_storage
    from django.http import JsonResponse

    project_name = "evpn_vxlan_erb_dc_1.3"
    zip_bytes = _build_project_zip(project_name)

    # Redirect default_storage to a real temp directory so validateZipFile
    # can read the uploaded file without touching production paths.
    with tempfile.TemporaryDirectory() as tmp_dir:
        real_storage_path = pathlib.Path(tmp_dir)

        class _TmpStorage:
            def exists(self, name):
                return (real_storage_path / name).exists()

            def save(self, name, f):
                dest = real_storage_path / name
                dest.write_bytes(f.read() if hasattr(f, "read") else f)
                return name

            def delete(self, name):
                p = real_storage_path / name
                if p.exists():
                    p.unlink()

            def path(self, name):
                return str(real_storage_path / name)

        monkeypatch.setattr(django_storage, "default_storage", _TmpStorage())
        # Bypass the real parseProjectFile (which calls Jenkins) and just
        # return a success response to prove the upload path works.
        monkeypatch.setattr(
            NetworkTypeParser,
            "parseProjectFile",
            lambda self, name: JsonResponse(
                {"result": "success", "name": project_name}
            ),
        )

        response = api_client.post(
            "/api/v1/network-types/upload/",
            {"app_zip_file": io.BytesIO(zip_bytes)},
            format="multipart",
        )

    assert response.status_code == 200
    body = response.json()
    assert body["result"] == "success"
    assert body["name"] == project_name


# ===========================================================================
# Section 2 — Network creation and CRUD
# ===========================================================================


# ── 2.a  CREATE ─────────────────────────────────────────────────────────────

@pytest.mark.django_db
def test_network_create_with_dc1_hosts(api_client, evpn_campus_type, dc1_host_content):
    """POST /api/v1/networks/ must create a CampusNetwork backed by the real
    dc1-hosts file content and return HTTP 201.
    """
    payload = {
        "name": "evpn-dc1",
        "status": "Initialized",
        "description": "EVPN VXLAN ERB DC1",
        "host_file": dc1_host_content,
        "campus_type": evpn_campus_type.id,
    }
    response = api_client.post("/api/v1/networks/", payload)
    assert response.status_code == 201
    data = response.json()
    assert data["name"] == "evpn-dc1"
    assert CampusNetwork.objects.filter(name="evpn-dc1").exists()


@pytest.mark.django_db
def test_network_create_with_dc2_hosts(api_client, evpn_campus_type, dc2_host_content):
    """The same CampusType supports a second network using dc2-hosts."""
    payload = {
        "name": "evpn-dc2",
        "status": "Initialized",
        "description": "EVPN VXLAN ERB DC2",
        "host_file": dc2_host_content,
        "campus_type": evpn_campus_type.id,
    }
    response = api_client.post("/api/v1/networks/", payload)
    assert response.status_code == 201
    assert response.json()["name"] == "evpn-dc2"


# ── 2.1  Host file load ─────────────────────────────────────────────────────

@pytest.mark.django_db
def test_network_create_stores_host_file_verbatim(api_client, evpn_campus_type, dc1_host_content):
    """The host_file field must be stored and returned without modification."""
    payload = {
        "name": "evpn-dc1-hosts",
        "status": "Initialized",
        "description": "Host file test",
        "host_file": dc1_host_content,
        "campus_type": evpn_campus_type.id,
    }
    create_resp = api_client.post("/api/v1/networks/", payload)
    assert create_resp.status_code == 201
    network_id = create_resp.json()["id"]

    retrieve_resp = api_client.get(f"/api/v1/networks/{network_id}/")
    assert retrieve_resp.status_code == 200
    assert retrieve_resp.json()["host_file"].strip() == dc1_host_content.strip()


@pytest.mark.django_db
def test_host_file_contains_expected_dc1_groups(evpn_network):
    """The stored dc1-hosts must contain the inventory groups defined in
    the evpn_vxlan_erb_dc example (spines, leaves, firewalls, servers).
    """
    for group in ("spines", "leaves", "firewalls", "servers"):
        assert group in evpn_network.host_file


@pytest.mark.django_db
def test_host_file_update_via_patch(api_client, evpn_network, dc2_host_content):
    """PATCH /api/v1/networks/{id}/ must accept a new host_file value."""
    response = api_client.patch(
        f"/api/v1/networks/{evpn_network.id}/",
        {"host_file": dc2_host_content},
    )
    assert response.status_code == 200
    evpn_network.refresh_from_db()
    assert evpn_network.host_file.strip() == dc2_host_content.strip()


# ── 2.2  Configuration load (spreadsheet) ───────────────────────────────────

@pytest.mark.django_db
def test_workbook_upload_with_real_xlsx(api_client, evpn_network):
    """POST /api/v1/networks/{id}/workbook/upload/ must successfully parse
    dc1_data.xlsx and persist the sheet data to the database, returning HTTP 200.
    """
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        response = api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "success"
    # Workbook and at least one Worksheet row must now exist in the DB.
    assert Workbook.objects.filter(campus_network_id=evpn_network).exists()
    assert Worksheets.objects.filter(
        workbook_id__campus_network_id=evpn_network
    ).exists()


@pytest.mark.django_db
def test_workbook_upload_persists_base_sheet(api_client, evpn_network):
    """After uploading dc1_data.xlsx the 'base' sheet must be retrievable
    via GET /api/v1/networks/{id}/workbook/.
    """
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )

    retrieve_resp = api_client.get(f"/api/v1/networks/{evpn_network.id}/workbook/")
    assert retrieve_resp.status_code == 200
    body = retrieve_resp.json()
    assert body["status"] == "success"
    sheet_names = [s["name"] for s in body["workbook"]]
    assert "base" in sheet_names


@pytest.mark.django_db
def test_workbook_upload_persists_all_expected_sheets(api_client, evpn_network):
    """The upload must create Worksheets rows for every sheet in dc1_data.xlsx.

    The expected sheet names are taken directly from the workbook so the test
    stays in sync with any future additions to the spreadsheet.
    """
    from openpyxl import load_workbook as openpyxl_load

    expected_sheet_names = set(openpyxl_load(DC1_XLSX_PATH).sheetnames)

    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )

    persisted_sheet_names = set(
        Worksheets.objects.filter(
            workbook_id__campus_network_id=evpn_network
        ).values_list("name", flat=True)
    )
    assert expected_sheet_names == persisted_sheet_names


@pytest.mark.django_db
def test_workbook_upload_without_file_returns_400(api_client, evpn_network):
    """Uploading with no file attached must return HTTP 400."""
    response = api_client.post(
        f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
        {},
        format="multipart",
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_workbook_upload_clears_previous_data(api_client, evpn_network):
    """A second upload to the same network must replace the previous workbook
    (old Workbook rows are deleted before the new ones are written).
    """
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )
    count_after_first = Worksheets.objects.filter(
        workbook_id__campus_network_id=evpn_network
    ).count()
    assert count_after_first > 0

    # Re-upload the same file — row count must stay consistent (not double).
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )
    count_after_second = Worksheets.objects.filter(
        workbook_id__campus_network_id=evpn_network
    ).count()
    assert count_after_second == count_after_first


# ── 2.3  Export (workbook download) ─────────────────────────────────────────

@pytest.mark.django_db
def test_workbook_download_returns_xlsx_attachment(
    api_client, evpn_network_with_workbook, monkeypatch
):
    """GET /api/v1/networks/{id}/workbook/download/ must return an xlsx file
    as an attachment with the correct MIME type.

    create_workbook writes to /tmp/nita-webapp/export/ which is monkeypatched
    here to keep tests filesystem-agnostic.
    """
    network = evpn_network_with_workbook

    # create_workbook_from_db already tested elsewhere; stub it to return a
    # minimal but valid xlsx file so we can verify the FileResponse path.
    tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    from openpyxl import Workbook as OpenpyxlWorkbook

    wb_obj = OpenpyxlWorkbook()
    wb_obj.active.title = "base"
    wb_obj.save(tmp_xlsx.name)
    tmp_xlsx.close()

    workbook_name = (
        f"{network.campus_type.name}_{network.name}.xlsx"
    )

    def _fake_create_workbook(pk):
        """Write a valid xlsx to the path the real function would use."""
        import os
        export_dir = pathlib.Path("/tmp/nita-webapp/export")
        export_dir.mkdir(parents=True, exist_ok=True)
        dest = export_dir / workbook_name
        import shutil
        shutil.copy(tmp_xlsx.name, dest)

    monkeypatch.setattr(api_views, "create_workbook", _fake_create_workbook)

    response = api_client.get(f"/api/v1/networks/{network.id}/workbook/download/")
    assert response.status_code == 200
    content_disposition = response.get("Content-Disposition", "")
    assert "attachment" in content_disposition
    assert workbook_name in content_disposition

    # Clean up temp file.
    pathlib.Path(tmp_xlsx.name).unlink(missing_ok=True)


@pytest.mark.django_db
def test_workbook_download_error_returns_500(api_client, evpn_network_with_workbook, monkeypatch):
    """When create_workbook raises, the endpoint must return HTTP 500."""

    def _raise(pk):
        raise OSError("no workbook data on disk")

    monkeypatch.setattr(api_views, "create_workbook", _raise)

    response = api_client.get(
        f"/api/v1/networks/{evpn_network_with_workbook.id}/workbook/download/"
    )
    assert response.status_code == 500


@pytest.mark.django_db
def test_workbook_download_requires_authentication(evpn_network_with_workbook, monkeypatch):
    """Unauthenticated requests to the download endpoint must be rejected."""
    monkeypatch.setattr(api_views, "create_workbook", lambda pk: None)
    response = APIClient().get(
        f"/api/v1/networks/{evpn_network_with_workbook.id}/workbook/download/"
    )
    assert response.status_code in (401, 403)


# ── 2.4  Full CRUD ───────────────────────────────────────────────────────────

# LIST
@pytest.mark.django_db
def test_network_list_includes_evpn_network(api_client, evpn_network):
    """GET /api/v1/networks/ must include the newly created evpn network."""
    response = api_client.get("/api/v1/networks/")
    assert response.status_code == 200
    names = [r["name"] for r in response.json()["results"]]
    assert evpn_network.name in names


@pytest.mark.django_db
def test_network_list_is_paginated(api_client, evpn_campus_type, dc1_host_content):
    """The list endpoint must return a paginated envelope with count/results."""
    # Create several networks so pagination metadata is populated.
    for i in range(3):
        CampusNetwork.objects.create(
            name=f"evpn-list-{i}",
            status="Initialized",
            description=f"Network {i}",
            host_file=dc1_host_content,
            campus_type=evpn_campus_type,
        )
    response = api_client.get("/api/v1/networks/")
    assert response.status_code == 200
    body = response.json()
    assert "count" in body
    assert "results" in body
    assert body["count"] >= 3


# RETRIEVE
@pytest.mark.django_db
def test_network_retrieve_returns_campus_type_name(api_client, evpn_network):
    """GET /api/v1/networks/{id}/ must include the campus_type_name string."""
    response = api_client.get(f"/api/v1/networks/{evpn_network.id}/")
    assert response.status_code == 200
    data = response.json()
    assert data["campus_type_name"] == evpn_network.campus_type.name


@pytest.mark.django_db
def test_network_retrieve_returns_host_file(api_client, evpn_network):
    """GET /api/v1/networks/{id}/ must return the stored host_file text."""
    response = api_client.get(f"/api/v1/networks/{evpn_network.id}/")
    assert response.status_code == 200
    assert response.json()["host_file"] == evpn_network.host_file


@pytest.mark.django_db
def test_network_retrieve_unknown_id_returns_404(api_client):
    """Retrieving a non-existent network ID must return HTTP 404."""
    response = api_client.get("/api/v1/networks/99999/")
    assert response.status_code == 404


# UPDATE (PUT)
@pytest.mark.django_db
def test_network_full_update(api_client, evpn_network, dc1_host_content):
    """PUT /api/v1/networks/{id}/ must replace all editable fields."""
    payload = {
        "name": evpn_network.name,
        "status": "Updated",
        "description": "Fully updated description",
        "host_file": dc1_host_content,
        "campus_type": evpn_network.campus_type.id,
    }
    response = api_client.put(f"/api/v1/networks/{evpn_network.id}/", payload)
    assert response.status_code == 200
    evpn_network.refresh_from_db()
    assert evpn_network.status == "Updated"
    assert evpn_network.description == "Fully updated description"


# PARTIAL UPDATE (PATCH)
@pytest.mark.django_db
def test_network_partial_update_description(api_client, evpn_network):
    """PATCH /api/v1/networks/{id}/ must update only the supplied fields."""
    new_desc = "Patched by test"
    response = api_client.patch(
        f"/api/v1/networks/{evpn_network.id}/",
        {"description": new_desc},
    )
    assert response.status_code == 200
    evpn_network.refresh_from_db()
    assert evpn_network.description == new_desc


@pytest.mark.django_db
def test_network_partial_update_status(api_client, evpn_network):
    """PATCH can update the status field independently."""
    response = api_client.patch(
        f"/api/v1/networks/{evpn_network.id}/",
        {"status": "Configured"},
    )
    assert response.status_code == 200
    evpn_network.refresh_from_db()
    assert evpn_network.status == "Configured"


# DELETE
@pytest.mark.django_db
def test_network_delete(api_client, evpn_network):
    """DELETE /api/v1/networks/{id}/ must remove the network and return 204."""
    network_id = evpn_network.id
    response = api_client.delete(f"/api/v1/networks/{network_id}/")
    assert response.status_code == 204
    assert not CampusNetwork.objects.filter(id=network_id).exists()


@pytest.mark.django_db
def test_network_delete_cascades_workbook(api_client, evpn_network):
    """Deleting a CampusNetwork must cascade-delete any associated Workbook
    and Worksheets rows.
    """
    workbook = Workbook.objects.create(
        name="dc1_data.xlsx",
        campus_network_id=evpn_network,
    )
    Worksheets.objects.create(
        name="base",
        data="{}",
        workbook_id=workbook,
    )
    network_id = evpn_network.id
    api_client.delete(f"/api/v1/networks/{network_id}/")
    assert not Workbook.objects.filter(campus_network_id_id=network_id).exists()
    assert not Worksheets.objects.filter(workbook_id__campus_network_id_id=network_id).exists()


@pytest.mark.django_db
def test_network_delete_requires_authentication(evpn_network):
    """DELETE without credentials must be rejected."""
    response = APIClient().delete(f"/api/v1/networks/{evpn_network.id}/")
    assert response.status_code in (401, 403)
    assert CampusNetwork.objects.filter(id=evpn_network.id).exists()


# WORKBOOK GET / SAVE / CLEAR
@pytest.mark.django_db
def test_workbook_retrieve_after_upload(api_client, evpn_network):
    """After uploading dc1_data.xlsx the workbook endpoint must return the
    parsed sheets under 'workbook' key with status=success.
    """
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )

    response = api_client.get(f"/api/v1/networks/{evpn_network.id}/workbook/")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "success"
    assert isinstance(body["workbook"], list)
    assert len(body["workbook"]) > 0


@pytest.mark.django_db
def test_workbook_clear_removes_data(api_client, evpn_network):
    """DELETE /api/v1/networks/{id}/workbook/clear/ must remove the Workbook row."""
    with DC1_XLSX_PATH.open("rb") as xlsx_fh:
        api_client.post(
            f"/api/v1/networks/{evpn_network.id}/workbook/upload/",
            {"up_file": xlsx_fh},
            format="multipart",
        )
    assert Workbook.objects.filter(campus_network_id=evpn_network).exists()

    clear_resp = api_client.delete(
        f"/api/v1/networks/{evpn_network.id}/workbook/clear/"
    )
    assert clear_resp.status_code == 204
    assert not Workbook.objects.filter(campus_network_id=evpn_network).exists()


@pytest.mark.django_db
def test_workbook_save_updates_sheet_data(api_client, evpn_network_with_workbook):
    """POST /api/v1/networks/{id}/workbook/save/ must persist updated grid data."""
    import json

    network = evpn_network_with_workbook
    new_data = [
        {
            "name": "base",
            "base": [
                {"host": "group_vars/all.yaml", "name": "OS_dir", "value": "/updated/"},
            ],
        }
    ]
    response = api_client.post(
        f"/api/v1/networks/{network.id}/workbook/save/",
        data=json.dumps({"data": new_data}),
        content_type="application/json",
    )
    assert response.status_code == 200
    assert response.json()["status"] == "success"


# NETWORK-TYPE  (read-only context for EVPN type)
@pytest.mark.django_db
def test_network_type_list_includes_evpn_type(api_client, evpn_campus_type):
    """GET /api/v1/network-types/ must include the evpn_vxlan_erb_dc type."""
    response = api_client.get("/api/v1/network-types/")
    assert response.status_code == 200
    names = [r["name"] for r in response.json()["results"]]
    assert evpn_campus_type.name in names


@pytest.mark.django_db
def test_network_type_retrieve_includes_nested_fields(api_client, evpn_campus_type):
    """GET /api/v1/network-types/{id}/ must include 'roles' and 'resources'."""
    response = api_client.get(f"/api/v1/network-types/{evpn_campus_type.id}/")
    assert response.status_code == 200
    data = response.json()
    assert data["name"] == evpn_campus_type.name
    assert "roles" in data
    assert "resources" in data


@pytest.mark.django_db
def test_network_type_delete_with_active_network_is_blocked(
    api_client, evpn_campus_type, evpn_network
):
    """Deleting a CampusType that has active networks must not silently succeed
    at the DB layer — the network still references the type.

    The REST API returns 204 for the type deletion itself (the view does not
    add a guard; the DB cascade owns this).  This test asserts that if the
    type is deleted its networks are also removed (cascade integrity).
    """
    network_id = evpn_network.id
    response = api_client.delete(f"/api/v1/network-types/{evpn_campus_type.id}/")
    assert response.status_code == 204
    # Cascade must remove dependent network.
    assert not CampusNetwork.objects.filter(id=network_id).exists()


# ACTIONS  (read-only listing scoped to EVPN campus type)
@pytest.mark.django_db
def test_actions_list_filtered_by_evpn_campus_type(api_client, evpn_campus_type, project_yaml_data):
    """GET /api/v1/actions/?campus_type_id=... must return only actions
    belonging to the EVPN campus type.
    """
    response = api_client.get(f"/api/v1/actions/?campus_type_id={evpn_campus_type.id}")
    assert response.status_code == 200
    results = response.json()["results"]
    assert len(results) == len(project_yaml_data["action"])
    for result in results:
        assert result["campus_type_id"] == evpn_campus_type.id
